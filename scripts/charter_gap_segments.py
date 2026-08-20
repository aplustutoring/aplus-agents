#!/usr/bin/env python3
"""Charter gap: full segmentation for the Monday launch (Roman 2026-08-14:
"split list as segmented as you can make them").

Reads the gap master (list 3104) + the workbook + each contact's charter
deals, computes per family:
  student_names / student_count  — every distinct student named across the
                                   family's charter deals since SINCE
  recency                        — hot (invoiced ≤90d) / winback / never
  personalized                   — last_tutor_name + student_first_name stamped
and builds one static list per cell of (recency × student-count × personalized):

  Charter 26/27 Monday - Hot - 1 student
  Charter 26/27 Monday - Hot - Multi student
  Charter 26/27 Monday - Win-back - 1 student
  Charter 26/27 Monday - Win-back - Multi student
  Charter 26/27 Monday - Never Started         (no tutor data by definition)
  Charter 26/27 Monday - No Lesson Data        (invoiced, but no completed lesson)

--write-props stamps student_names/student_count onto the contacts (UPDATE-only
import, needs crm.import). --build-lists creates/refreshes the lists.
Read-only otherwise. Deal-based end to end (no invoice-status logic).
"""

import argparse
import io
import json
import os
import re
import sys
import time
from datetime import date
from pathlib import Path

import requests
from openpyxl import load_workbook

try:
    from dotenv import load_dotenv
    _here = Path(__file__).resolve().parent.parent
    load_dotenv(_here / ".env", override=False)
    # worktrees have no .env — fall back to the main checkout's
    if "/.claude/worktrees/" in str(_here):
        load_dotenv(Path(str(_here).split("/.claude/worktrees/")[0]) / ".env", override=False)
except ImportError:
    pass

HUBSPOT_TOKEN = os.getenv("HUBSPOT_PRIVATE_APP_TOKEN", "") or os.getenv("HUBSPOT_API_KEY", "")
HS_BASE = "https://api.hubapi.com"
H = {"Authorization": f"Bearer {HUBSPOT_TOKEN}", "Content-Type": "application/json"}

GAP_LIST_ID = 3104
SINCE = "2025-08-01"
HOT_DAYS = 90
CHARTER_PIPELINES = {"907748", "72281989", "88841552", "5119061", "1066195"}
LIST_PREFIX = "Charter 26/27 Monday - "
NOT_STUDENT_TOKENS = {"a", "summer", "level", "ilead", "charter"}


def hs(method, path, **kw):
    for _ in range(5):
        r = requests.request(method, f"{HS_BASE}{path}", headers=H, timeout=30, **kw)
        if r.status_code == 429:
            time.sleep(int(r.headers.get("Retry-After", 10)) or 10)
            continue
        r.raise_for_status()
        return r.json() if r.text else {}
    r.raise_for_status()


def list_members(list_id):
    ids, after = [], None
    while True:
        j = hs("GET", f"/crm/v3/lists/{list_id}/memberships?limit=250" + (f"&after={after}" if after else ""))
        ids += [str(m["recordId"]) for m in j.get("results", [])]
        after = (j.get("paging", {}).get("next") or {}).get("after")
        if not after:
            return ids


def batch_read(object_type, ids, props):
    out = {}
    for i in range(0, len(ids), 100):
        j = hs("POST", f"/crm/v3/objects/{object_type}/batch/read",
               json={"inputs": [{"id": c} for c in ids[i:i + 100]], "properties": props})
        for row in j.get("results", []):
            out[str(row["id"])] = row.get("properties", {})
    return out


def contact_deals(ids):
    out = {}
    for i in range(0, len(ids), 100):
        j = hs("POST", "/crm/v4/associations/contact/deal/batch/read",
               json={"inputs": [{"id": c} for c in ids[i:i + 100]]})
        for row in j.get("results", []):
            out[str(row["from"]["id"])] = [str(t["toObjectId"]) for t in row.get("to", [])]
        time.sleep(0.1)
    return out


def students_from_dealnames(names, parent_first):
    """Distinct student first names from 'Parent - Student - School N - YY/YY'
    deal titles. Case-insensitive dedupe; drops the parent's own name and
    prefix tokens ('A -', 'Summer 2025 -')."""
    seen = {}
    for n in names:
        parts = [p.strip() for p in n.split(" - ")]
        if len(parts) < 3:
            continue
        cand = parts[1].split(" ")[0].strip(" .,")
        key = cand.lower()
        if not cand or key == parent_first or key in NOT_STUDENT_TOKENS or not re.match(r"^[A-Za-zÀ-ÿ'\-]+$", cand):
            continue
        seen.setdefault(key, cand)
    return sorted(seen.values(), key=str.lower)


def prose(names):
    if not names:
        return ""
    if len(names) == 1:
        return names[0]
    return ", ".join(names[:-1]) + " & " + names[-1]


def import_props(rows):
    """rows: [(email, student_names, student_count)] → UPDATE-only import."""
    import csv
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["Email", "Student Names", "Student Count"])
    for em, sn, sc in rows:
        w.writerow([em, sn, sc])
    req = {"name": f"Charter gap student names ({date.today().isoformat()})",
           "importOperations": {"0-1": "UPDATE"},
           "files": [{"fileName": "gap_students.csv", "fileFormat": "CSV",
                      "fileImportPage": {"hasHeader": True, "columnMappings": [
                          {"columnObjectTypeId": "0-1", "columnName": "Email",
                           "propertyName": "email", "columnType": "HUBSPOT_ALTERNATE_ID"},
                          {"columnObjectTypeId": "0-1", "columnName": "Student Names",
                           "propertyName": "student_names"},
                          {"columnObjectTypeId": "0-1", "columnName": "Student Count",
                           "propertyName": "student_count"}]}}]}
    r = requests.post(f"{HS_BASE}/crm/v3/imports", headers={"Authorization": f"Bearer {HUBSPOT_TOKEN}"},
                      files={"files": ("gap_students.csv", io.BytesIO(buf.getvalue().encode()), "text/csv")},
                      data={"importRequest": json.dumps(req)}, timeout=60)
    if r.status_code >= 400:
        sys.exit(f"import failed {r.status_code}: {r.text[:300]}")
    imp = r.json().get("importId") or r.json().get("id")
    state = ""
    for _ in range(30):
        time.sleep(10)
        state = hs("GET", f"/crm/v3/imports/{imp}").get("state", "")
        if state in ("DONE", "FAILED", "CANCELED", "DEFERRED"):
            break
    print(f"    import {imp}: {state}")


def upsert_list(name, ids):
    j = hs("POST", "/crm/v3/lists/search", json={"query": name, "count": 10})
    ex = [l for l in j.get("lists", []) if l.get("name") == name]
    if ex:
        lid = ex[0]["listId"]
        # refresh: remove everyone not in ids, add missing
        current = set(list_members(lid))
        rm = sorted(current - set(ids))
        for i in range(0, len(rm), 250):
            hs("PUT", f"/crm/v3/lists/{lid}/memberships/remove", json=rm[i:i + 250])
    else:
        lid = hs("POST", "/crm/v3/lists", json={"name": name, "objectTypeId": "0-1",
                                                "processingType": "MANUAL"})["list"]["listId"]
    for i in range(0, len(ids), 250):
        hs("PUT", f"/crm/v3/lists/{lid}/memberships/add", json=ids[i:i + 250])
    return lid


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--workbook", default=str(Path.home() / "Desktop" / "charter_gap_analysis.xlsx"))
    ap.add_argument("--write-props", action="store_true")
    ap.add_argument("--build-lists", action="store_true")
    args = ap.parse_args()

    # recency from the workbook (deal-based gap set; invoice date only for hot/winback)
    wb = load_workbook(args.workbook)
    recency = {}
    for r in wb["GAP - Priority Targets"].iter_rows(min_row=2, values_only=True):
        recency[r[12].rstrip("/").rsplit("/", 1)[-1]] = "hot" if r[8] == "Hot" else "winback"
    for r in wb["GAP - Deal But Never Invoiced"].iter_rows(min_row=2, values_only=True):
        recency[r[7].rstrip("/").rsplit("/", 1)[-1]] = "never"

    ids = list_members(GAP_LIST_ID)
    contacts = batch_read("contacts", ids, ["firstname", "lastname", "email",
                                            "last_tutor_name", "student_first_name"])
    assoc = contact_deals(ids)
    deal_ids = sorted({d for ds in assoc.values() for d in ds})
    deals = batch_read("deals", deal_ids, ["dealname", "pipeline", "createdate"])
    print(f"gap contacts {len(ids)}, charter deals in window "
          f"{sum(1 for d in deals.values() if d.get('pipeline') in CHARTER_PIPELINES and (d.get('createdate') or '') >= SINCE)}")

    rows, cells = [], {}
    for cid in ids:
        p = contacts.get(cid, {})
        names = [deals[d]["dealname"] for d in assoc.get(cid, [])
                 if d in deals and deals[d].get("pipeline") in CHARTER_PIPELINES
                 and (deals[d].get("createdate") or "") >= SINCE and deals[d].get("dealname")]
        studs = students_from_dealnames(names, (p.get("firstname") or "").strip().lower())
        # fall back to the stamped lesson student if deal names gave nothing
        if not studs and (p.get("student_first_name") or "").strip():
            studs = [p["student_first_name"].strip()]
        rows.append((cid, (p.get("email") or "").strip(), prose(studs), len(studs)))
        rec = recency.get(cid, "never")
        personalized = bool((p.get("last_tutor_name") or "").strip() and (p.get("student_first_name") or "").strip())
        if rec == "never":
            cell = "Never Started"
        elif not personalized:
            cell = "No Lesson Data"
        else:
            cell = f"{'Hot' if rec == 'hot' else 'Win-back'} - {'1 student' if len(studs) <= 1 else 'Multi student'}"
        cells.setdefault(cell, []).append(cid)

    print("\nSEGMENTS:")
    for cell in ["Hot - 1 student", "Hot - Multi student", "Win-back - 1 student",
                 "Win-back - Multi student", "Never Started", "No Lesson Data"]:
        print(f"  {cell:<26} {len(cells.get(cell, [])):>4}")
    multi = [r for r in rows if r[3] > 1]
    print(f"\nmulti-student families: {len(multi)}; sample: "
          + "; ".join(f"{contacts[r[0]].get('firstname')} {contacts[r[0]].get('lastname')}: {r[2]}" for r in multi[:5]))

    if args.write_props:
        todo = [(em, sn, sc) for _, em, sn, sc in rows if em and sn]
        print(f"\nstamping student_names/student_count on {len(todo)} contacts...")
        import_props(todo)
        chk = batch_read("contacts", [r[0] for r in rows[:100]], ["student_names"])
        print(f"    verify sample: {sum(1 for v in chk.values() if (v.get('student_names') or '').strip())}/{len(chk)} stamped")

    if args.build_lists:
        print("\nLISTS:")
        for cell, members in cells.items():
            lid = upsert_list(LIST_PREFIX + cell, members)
            print(f"  {lid} | {LIST_PREFIX}{cell}: {len(members)}")


if __name__ == "__main__":
    main()
