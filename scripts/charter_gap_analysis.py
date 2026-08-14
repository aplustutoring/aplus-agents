#!/usr/bin/env python3
"""Charter renewal gap analysis — read-only by default against HubSpot + Teachworks.

Finds charter families with a 25/26 deal but no 26/27 renewal, enriched with
Teachworks invoice history AND each family's most recent completed lesson
(tutor name + student first name), and writes a 4-tab xlsx:

  1. GAP - Priority Targets ....... gap families WITH invoices (Hot / Win-back)
  2. GAP - Deal But Never Invoiced  gap families with no Teachworks match
  3. Renewed 26-27 ................ welcome-back list
  4. Mismatches ................... TW-invoiced families with no charter deal
                                    since Aug 2025 (data hygiene queue)

Charter pipelines: 907748 (Traditional Vendor Funds), 72281989 (Terri iLead),
88841552 (Amy iLead), 5119061 (IEM Inc.), 1066195 (CFGC).

A family is RENEWED if any of its charter deals was created on/after
2026-08-01 OR has "26/27" in the dealname; otherwise it is a GAP family.
School staff are excluded by email domain (student.* subdomains stay family).

The ONLY write is opt-in: --write-props stamps last_tutor_name +
student_first_name (declared in ops/hubspot-schema/properties.yml) onto
gap contacts on static list 3104 via an UPDATE-only email-keyed import
(needs the crm.import scope; creates no contacts, builds no lists).

Auth (env, or repo-root .env via python-dotenv):
  HUBSPOT_PRIVATE_APP_TOKEN (or HUBSPOT_API_KEY)
  TEACHWORKS_TOKEN (or TEACHWORKS_TOKEN_ONLINE)   — online account
  TEACHWORKS_TOKEN_INPERSON                        — in-person account (optional)

Usage:
  python3 scripts/charter_gap_analysis.py [--out PATH] [--write-props]
Defaults to ~/Desktop/charter_gap_analysis.xlsx.

One-off built 2026-08-13 (Roman); structured for later scheduling.
"""

import argparse
import os
import re
import sys
import time
from datetime import date, datetime
from pathlib import Path

import requests
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

load_dotenv(Path(__file__).resolve().parent.parent / ".env", override=False)

HUBSPOT_TOKEN = os.getenv("HUBSPOT_PRIVATE_APP_TOKEN", "") or os.getenv("HUBSPOT_API_KEY", "")
TW_TOKENS = {}  # account label -> token
if os.getenv("TEACHWORKS_TOKEN", "") or os.getenv("TEACHWORKS_TOKEN_ONLINE", ""):
    TW_TOKENS["online"] = os.getenv("TEACHWORKS_TOKEN", "") or os.getenv("TEACHWORKS_TOKEN_ONLINE", "")
if os.getenv("TEACHWORKS_TOKEN_INPERSON", ""):
    TW_TOKENS["in_person"] = os.getenv("TEACHWORKS_TOKEN_INPERSON", "")

HS_BASE = "https://api.hubapi.com"
TW_BASE = "https://api.teachworks.com/v1"

SINCE = "2025-08-01"            # window start for deals and invoices
RENEWAL_CUTOFF = "2026-08-01"   # deals created on/after this date = 26/27 renewal
                                # (was 2026-06-01; tightened 2026-08-13 — late
                                # spring 25/26 POs were counting as renewals)
RENEWAL_NAME_RE = re.compile(r"26\s*[/\-]\s*27")
HOT_DAYS = 90                   # last invoice within this many days = "Hot"

CHARTER_PIPELINES = {
    "907748": "Charter Schools - Traditional Vendor Funds",
    "72281989": "Terri iLead Level Up",
    "88841552": "Amy iLead Level Up",
    "5119061": "IEM Inc.",
    "1066195": "CFGC",
}

# School-staff email domains (exact match, or non-student.* subdomain).
SCHOOL_DOMAINS = {
    "ileadexploration.org", "ileadav.org", "ileadlancaster.org", "ieminc.org",
    "viedu.org", "gormanlc.org", "eliteacademic.com", "heartlandcharterschool.com",
    "compasscharters.org", "pacificcharters.org", "pacificcoastacademy.org",
    "granitemountainschool.com", "heartwoodcharterschool.org",
    "theblueridgeacademy.com", "hcs.k12.ca.us", "forestcharter.com",
    "taylion.com", "suncoastprep.org", "sageoak.education",
}

# Sanity-check expectations (Roman, as of 2026-08-13). Warn if off by >30%.
EXPECT = {"families": 439, "renewed": 11, "gap": 428}

HUBSPOT_RECORD_URL = "https://app.hubspot.com/contacts/6312752/record/0-1/{id}"
GAP_LIST_ID = 3104              # "Charter 26/27 Gap Families" static list


# ─────────────────────────────────────────────
# HTTP helpers (retry on rate limit)
# ─────────────────────────────────────────────

def hs_request(method, path, **kwargs):
    headers = {"Authorization": f"Bearer {HUBSPOT_TOKEN}", "Content-Type": "application/json"}
    for attempt in range(5):
        r = requests.request(method, f"{HS_BASE}{path}", headers=headers, timeout=30, **kwargs)
        if r.status_code == 429:
            wait = int(r.headers.get("Retry-After", 10)) or 10
            print(f"      ⏳ HubSpot rate limit, retrying in {wait}s...")
            time.sleep(wait)
            continue
        r.raise_for_status()
        return r.json()
    r.raise_for_status()


def tw_get(endpoint, params=None, token=None):
    """Paginated GET against one Teachworks account (max 80/page, 403 backoff)."""
    headers = {"Authorization": f"Token token={token}", "Content-Type": "application/json"}
    params = dict(params or {})
    params["per_page"] = 80
    params["page"] = 1
    results = []
    while True:
        for attempt in range(3):
            r = requests.get(f"{TW_BASE}/{endpoint}", headers=headers, params=params, timeout=30)
            if r.status_code == 403:
                wait = 5 * (attempt + 1)
                print(f"      ⏳ Teachworks rate limit, retrying in {wait}s...")
                time.sleep(wait)
                continue
            r.raise_for_status()
            break
        else:
            r.raise_for_status()
        data = r.json()
        if not data:
            break
        results.extend(data)
        if len(data) < 80:
            break
        params["page"] += 1
    return results


# ─────────────────────────────────────────────
# STEP 1 — HubSpot: charter deals + contacts
# ─────────────────────────────────────────────

def fetch_charter_deals():
    """All deals created SINCE→now in the charter pipelines."""
    since_ms = int(datetime.fromisoformat(SINCE).timestamp() * 1000)
    deals = []
    after = None
    while True:
        body = {
            "filterGroups": [{"filters": [
                {"propertyName": "pipeline", "operator": "IN",
                 "values": list(CHARTER_PIPELINES)},
                {"propertyName": "createdate", "operator": "GTE", "value": str(since_ms)},
            ]}],
            "properties": ["dealname", "pipeline", "dealstage", "createdate", "amount"],
            "sorts": [{"propertyName": "createdate", "direction": "ASCENDING"}],
            "limit": 200,
        }
        if after:
            body["after"] = after
        page = hs_request("POST", "/crm/v3/objects/deals/search", json=body)
        deals.extend(page.get("results", []))
        after = (page.get("paging", {}).get("next") or {}).get("after")
        if not after:
            break
    if len(deals) >= 10000:
        print("  ⚠️  Hit the 10k search cap — results may be truncated!")
    return deals


def fetch_deal_contacts(deal_ids):
    """deal id -> [contact ids] via v4 batch association read."""
    assoc = {}
    for i in range(0, len(deal_ids), 100):
        chunk = deal_ids[i:i + 100]
        page = hs_request("POST", "/crm/v4/associations/deal/contact/batch/read",
                          json={"inputs": [{"id": d} for d in chunk]})
        for row in page.get("results", []):
            frm = str(row["from"]["id"])
            assoc[frm] = [str(t["toObjectId"]) for t in row.get("to", [])]
    return assoc


def fetch_contacts(contact_ids):
    """contact id -> {firstname, lastname, email, phone}."""
    out = {}
    ids = sorted(set(contact_ids))
    for i in range(0, len(ids), 100):
        chunk = ids[i:i + 100]
        page = hs_request("POST", "/crm/v3/objects/contacts/batch/read",
                          json={"inputs": [{"id": c} for c in chunk],
                                "properties": ["firstname", "lastname", "email", "phone"]})
        for row in page.get("results", []):
            p = row.get("properties", {})
            out[str(row["id"])] = {
                "id": str(row["id"]),
                "first": (p.get("firstname") or "").strip(),
                "last": (p.get("lastname") or "").strip(),
                "email": (p.get("email") or "").strip().lower(),
                "phone": (p.get("phone") or "").strip(),
            }
    return out


def is_school_staff(email):
    """Exact school domain, or a non-student.* subdomain of one."""
    if "@" not in email:
        return False
    domain = email.rsplit("@", 1)[1].lower()
    if domain in SCHOOL_DOMAINS:
        return True
    for d in SCHOOL_DOMAINS:
        if domain.endswith("." + d):
            sub = domain[: -len(d) - 1]
            return not sub.startswith("student")
    return False


def deal_is_renewal(deal):
    p = deal.get("properties", {})
    created = (p.get("createdate") or "")[:10]
    if created >= RENEWAL_CUTOFF:
        return True
    return bool(RENEWAL_NAME_RE.search(p.get("dealname") or ""))


# ─────────────────────────────────────────────
# STEP 2 — Teachworks: invoices per family
# ─────────────────────────────────────────────

def norm_name(first, last):
    return re.sub(r"\s+", " ", f"{first} {last}".strip().lower())


def fetch_tw_families():
    """Aggregate non-void invoices SINCE→today per Teachworks customer,
    across both accounts. Returns list of family dicts."""
    today = date.today().isoformat()
    fams = {}  # (acct, customer_id) -> agg
    for acct, token in TW_TOKENS.items():
        print(f"    [{acct}] pulling invoices {SINCE} → {today}...")
        invoices = tw_get("invoices", {"date[gte]": SINCE, "date[lte]": today}, token=token)
        void = sum(1 for i in invoices if (i.get("status") or "") == "Void")
        print(f"    [{acct}] {len(invoices)} invoices ({void} void, excluded)")
        needed = {i.get("customer_id") for i in invoices if i.get("customer_id")}
        print(f"    [{acct}] pulling customer records for {len(needed)} customers...")
        customers = {c["id"]: c for c in tw_get("customers", token=token)}
        missing_cust = needed - set(customers)
        if missing_cust:
            print(f"    ⚠️  [{acct}] {len(missing_cust)} invoice customer_ids not in customer list")
        for inv in invoices:
            if (inv.get("status") or "") == "Void":
                continue
            cid = inv.get("customer_id")
            cust = customers.get(cid, {})
            key = (acct, cid)
            f = fams.setdefault(key, {
                "account": acct,
                # customer record first; invoice-level name fields as fallback
                # (dupe/archived customers can be absent from the customer list)
                "first": (cust.get("first_name") or inv.get("customer_first_name") or "").strip(),
                "last": (cust.get("last_name") or inv.get("customer_last_name") or "").strip(),
                "email": (cust.get("email") or "").strip().lower(),
                "customer_ids": [(acct, cid)],
                "last_invoice_date": "",
                "total_invoiced": 0.0,
                "invoice_count": 0,
            })
            f["invoice_count"] += 1
            try:
                f["total_invoiced"] += float(inv.get("total") or inv.get("amount") or 0)
            except (TypeError, ValueError):
                pass
            d = str(inv.get("date") or "")[:10]
            if d > f["last_invoice_date"]:
                f["last_invoice_date"] = d
    # Merge duplicate customers across accounts / records on email (fallback name)
    merged = {}
    for f in fams.values():
        key = f["email"] or ("name:" + norm_name(f["first"], f["last"]))
        if key in ("", "name:"):
            key = f"anon:{id(f)}"
        m = merged.get(key)
        if not m:
            merged[key] = f
        else:
            m["invoice_count"] += f["invoice_count"]
            m["total_invoiced"] += f["total_invoiced"]
            m["last_invoice_date"] = max(m["last_invoice_date"], f["last_invoice_date"])
            m["customer_ids"].extend(f["customer_ids"])
            if f["account"] not in m["account"]:
                m["account"] += f"+{f['account']}"
    return list(merged.values())


def fetch_last_lessons(families):
    """For each TW family, the most recent completed lesson since SINCE:
    {"lesson_date", "tutor", "student_first"} stored on the family dict as
    f["last_lesson"]. Uses the proven per-customer students -> per-student
    lessons query pattern (same as email/src/teachworks_client.py)."""
    enriched = 0
    for f in families:
        best = None
        for acct, cid in f.get("customer_ids", []):
            token = TW_TOKENS.get(acct)
            if not token or not cid:
                continue
            for s in tw_get("students", {"customer_id": cid}, token=token):
                for l in tw_get("lessons", {"student_id": s["id"],
                                            "from_date[gte]": SINCE}, token=token):
                    status = str(l.get("status") or "").lower()
                    if "attend" not in status and "complete" not in status:
                        continue
                    d = str(l.get("from_date") or "")[:10]
                    if not best or d > best["lesson_date"]:
                        best = {"lesson_date": d,
                                "tutor": (l.get("employee_name") or "").strip(),
                                "student_first": (s.get("first_name") or "").strip()}
        if best:
            f["last_lesson"] = best
            enriched += 1
    return enriched


# ─────────────────────────────────────────────
# STEP 4 — xlsx output
# ─────────────────────────────────────────────

def write_sheet(wb, title, headers, rows, link_col=None):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
    for row in rows:
        ws.append(row)
    if link_col is not None:
        for r in range(2, len(rows) + 2):
            cell = ws.cell(row=r, column=link_col)
            if cell.value:
                cell.hyperlink = cell.value
                cell.font = Font(color="0563C1", underline="single")
    for i, h in enumerate(headers, 1):
        width = max([len(str(h))] + [len(str(r[i - 1] or "")) for r in rows[:200]]) + 2
        ws.column_dimensions[get_column_letter(i)].width = min(width, 50)
    ws.freeze_panes = "A2"
    return ws


def months_since(iso_date):
    if not iso_date:
        return ""
    days = (date.today() - date.fromisoformat(iso_date)).days
    return round(days / 30.44, 1)


def write_tutor_props(rows):
    """Stamp last_tutor_name + student_first_name onto list-3104 gap contacts
    via an UPDATE-only email-keyed import (crm.import scope; creates nothing).
    rows: [(contact_id, email, tutor, student_first)]."""
    import csv
    import io
    import json

    members = set()
    after = None
    while True:
        path = f"/crm/v3/lists/{GAP_LIST_ID}/memberships?limit=250"
        if after:
            path += f"&after={after}"
        page = hs_request("GET", path)
        members |= {str(m["recordId"]) for m in page.get("results", [])}
        after = (page.get("paging", {}).get("next") or {}).get("after")
        if not after:
            break
    todo = [r for r in rows if r[0] in members and r[1] and r[2]]
    skipped = len(rows) - len(todo)
    print(f"    list {GAP_LIST_ID}: {len(members)} members; writing {len(todo)} "
          f"contacts ({skipped} skipped: off-list, no email, or no tutor)")
    if not todo:
        return

    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["Email", "Last Tutor Name", "Student First Name"])
    for _, email, tutor, student in todo:
        w.writerow([email, tutor, student])
    import_request = {
        "name": f"Charter gap tutor enrichment ({date.today().isoformat()})",
        "importOperations": {"0-1": "UPDATE"},
        "files": [{
            "fileName": "gap_tutor_props.csv",
            "fileFormat": "CSV",
            "fileImportPage": {
                "hasHeader": True,
                "columnMappings": [
                    {"columnObjectTypeId": "0-1", "columnName": "Email",
                     "propertyName": "email", "columnType": "HUBSPOT_ALTERNATE_ID"},
                    {"columnObjectTypeId": "0-1", "columnName": "Last Tutor Name",
                     "propertyName": "last_tutor_name"},
                    {"columnObjectTypeId": "0-1", "columnName": "Student First Name",
                     "propertyName": "student_first_name"},
                ],
            },
        }],
    }
    r = requests.post(f"{HS_BASE}/crm/v3/imports",
                      headers={"Authorization": f"Bearer {HUBSPOT_TOKEN}"},
                      files={"files": ("gap_tutor_props.csv",
                                       io.BytesIO(buf.getvalue().encode()), "text/csv")},
                      data={"importRequest": json.dumps(import_request)}, timeout=60)
    if r.status_code >= 400:
        print(f"    ❌ import failed {r.status_code}: {r.text[:400]}")
        r.raise_for_status()
    imp_id = r.json().get("importId") or r.json().get("id")
    print(f"    import {imp_id} started...")
    state = ""
    for _ in range(30):
        time.sleep(10)
        state = hs_request("GET", f"/crm/v3/imports/{imp_id}").get("state", "")
        if state in ("DONE", "FAILED", "CANCELED", "DEFERRED"):
            break
    print(f"    import state: {state}")

    sample = [r[0] for r in todo[:100]]
    check = hs_request("POST", "/crm/v3/objects/contacts/batch/read",
                       json={"inputs": [{"id": c} for c in sample],
                             "properties": ["last_tutor_name", "student_first_name"]})
    stamped = sum(1 for row in check.get("results", [])
                  if (row.get("properties", {}).get("last_tutor_name") or "").strip())
    print(f"    verify sample: {stamped}/{len(sample)} have last_tutor_name")


def main():
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("--out", default=str(Path.home() / "Desktop" / "charter_gap_analysis.xlsx"))
    ap.add_argument("--write-props", action="store_true",
                    help=f"stamp last_tutor_name/student_first_name onto list-{GAP_LIST_ID} "
                         "contacts via UPDATE-only import (the script's only write)")
    args = ap.parse_args()

    missing = []
    if not HUBSPOT_TOKEN:
        missing.append("HUBSPOT_PRIVATE_APP_TOKEN")
    if not TW_TOKENS:
        missing.append("TEACHWORKS_TOKEN (or TEACHWORKS_TOKEN_ONLINE)")
    if missing:
        sys.exit(f"Missing env: {', '.join(missing)}")

    # ── Step 1: HubSpot ──
    print("📥 STEP 1 — HubSpot charter deals...")
    deals = fetch_charter_deals()
    print(f"    {len(deals)} deals since {SINCE} across {len(CHARTER_PIPELINES)} pipelines")
    deal_ids = [str(d["id"]) for d in deals]
    assoc = fetch_deal_contacts(deal_ids)
    contacts = fetch_contacts([c for ids in assoc.values() for c in ids])
    print(f"    {len(contacts)} associated contacts")

    families = {}   # contact id -> {contact, deals}
    staff_excluded = 0
    for deal in deals:
        for cid in assoc.get(str(deal["id"]), []):
            c = contacts.get(cid)
            if not c:
                continue
            if is_school_staff(c["email"]):
                staff_excluded += 1
                continue
            fam = families.setdefault(cid, {"contact": c, "deals": []})
            fam["deals"].append(deal)
    staff_ids = {cid for cid in contacts
                 if is_school_staff(contacts[cid]["email"])}
    print(f"    {len(families)} family contacts ({len(staff_ids)} school-staff contacts excluded)")

    renewed, gap = {}, {}
    for cid, fam in families.items():
        (renewed if any(deal_is_renewal(d) for d in fam["deals"]) else gap)[cid] = fam
    print(f"    RENEWED: {len(renewed)}   GAP: {len(gap)}")
    for k, got in (("families", len(families)), ("renewed", len(renewed)), ("gap", len(gap))):
        exp = EXPECT[k]
        if exp and abs(got - exp) / exp > 0.30:
            print(f"    ⚠️  SANITY CHECK: {k}={got}, expected ~{exp} — differs wildly, verify before using!")

    # ── Step 2: Teachworks ──
    print("📥 STEP 2 — Teachworks invoices...")
    tw_fams = fetch_tw_families()
    print(f"    {len(tw_fams)} Teachworks families with invoices since {SINCE}")

    # ── Step 3: merge ──
    print("🔗 STEP 3 — matching...")
    tw_by_email = {f["email"]: f for f in tw_fams if f["email"]}
    tw_by_name = {}
    for f in tw_fams:
        n = norm_name(f["first"], f["last"])
        if n:
            tw_by_name.setdefault(n, f)

    matched_tw = set()

    def tw_match(contact):
        f = tw_by_email.get(contact["email"]) if contact["email"] else None
        how = "email"
        if not f:
            f = tw_by_name.get(norm_name(contact["first"], contact["last"]))
            how = "name"
        if f:
            matched_tw.add(id(f))
        return f, (how if f else None)

    name_matches = 0
    gap_with_inv, gap_no_inv = [], []
    for cid, fam in gap.items():
        f, how = tw_match(fam["contact"])
        if how == "name":
            name_matches += 1
        (gap_with_inv if f else gap_no_inv).append((fam, f))
    renewed_rows = []
    for cid, fam in renewed.items():
        f, how = tw_match(fam["contact"])
        renewed_rows.append((fam, f))
    unmatched_tw = [f for f in tw_fams if id(f) not in matched_tw]
    print(f"    gap matched to TW: {len(gap_with_inv)} ({name_matches} by name), "
          f"unmatched gap: {len(gap_no_inv)}")
    print(f"    TW families with no charter-deal match: {len(unmatched_tw)}")

    # ── Step 3b: last completed lesson per matched gap family ──
    print("📥 STEP 3b — last completed lesson (tutor + student) for matched gap families...")
    seen_ids = set()
    to_enrich = []
    for _, f in gap_with_inv:
        if id(f) not in seen_ids:
            seen_ids.add(id(f))
            to_enrich.append(f)
    enriched = fetch_last_lessons(to_enrich)
    with_tutor = sum(1 for _, f in gap_with_inv
                     if f.get("last_lesson", {}).get("tutor"))
    total_gap = len(gap)
    print(f"    MATCH RATE: {with_tutor}/{total_gap} gap families got a tutor name "
          f"({with_tutor / total_gap * 100:.0f}%) — "
          f"{enriched} of {len(to_enrich)} matched TW families had a completed lesson")

    # ── Step 4: xlsx ──
    print(f"📤 STEP 4 — writing {args.out}...")
    wb = Workbook()
    wb.remove(wb.active)

    hdr1 = ["Last", "First", "Email", "Phone", "Last Invoice", "Total Invoiced",
            "Invoices", "Months Since Last Invoice", "Segment",
            "Last Tutor", "Student First", "Last Lesson", "HubSpot"]
    rows1 = []
    for fam, f in sorted(gap_with_inv, key=lambda x: -x[1]["total_invoiced"]):
        c = fam["contact"]
        seg = "Hot" if f["last_invoice_date"] and \
            (date.today() - date.fromisoformat(f["last_invoice_date"])).days <= HOT_DAYS \
            else "Win-back"
        ll = f.get("last_lesson", {})
        rows1.append([c["last"], c["first"], c["email"], c["phone"],
                      f["last_invoice_date"], round(f["total_invoiced"], 2),
                      f["invoice_count"], months_since(f["last_invoice_date"]),
                      seg, ll.get("tutor", ""), ll.get("student_first", ""),
                      ll.get("lesson_date", ""), HUBSPOT_RECORD_URL.format(id=c["id"])])
    write_sheet(wb, "GAP - Priority Targets", hdr1, rows1, link_col=len(hdr1))

    hdr2 = ["Last", "First", "Email", "Phone", "Deals", "Most Recent Deal",
            "Deal Names", "HubSpot"]
    rows2 = []
    for fam, _ in sorted(gap_no_inv, key=lambda x: (x[0]["contact"]["last"], x[0]["contact"]["first"])):
        c = fam["contact"]
        created = max((d["properties"].get("createdate") or "")[:10] for d in fam["deals"])
        names = "; ".join(sorted({d["properties"].get("dealname") or "" for d in fam["deals"]}))
        rows2.append([c["last"], c["first"], c["email"], c["phone"], len(fam["deals"]),
                      created, names[:200], HUBSPOT_RECORD_URL.format(id=c["id"])])
    write_sheet(wb, "GAP - Deal But Never Invoiced", hdr2, rows2, link_col=len(hdr2))

    hdr3 = ["Last", "First", "Email", "Phone", "Renewal Deal(s)", "Last Invoice",
            "Total Invoiced", "HubSpot"]
    rows3 = []
    for fam, f in sorted(renewed_rows, key=lambda x: (x[0]["contact"]["last"], x[0]["contact"]["first"])):
        c = fam["contact"]
        names = "; ".join(sorted({d["properties"].get("dealname") or ""
                                  for d in fam["deals"] if deal_is_renewal(d)}))
        rows3.append([c["last"], c["first"], c["email"], c["phone"], names[:200],
                      f["last_invoice_date"] if f else "",
                      round(f["total_invoiced"], 2) if f else "",
                      HUBSPOT_RECORD_URL.format(id=c["id"])])
    write_sheet(wb, "Renewed 26-27", hdr3, rows3, link_col=len(hdr3))

    hdr4 = ["Last", "First", "Email", "TW Account", "Last Invoice",
            "Total Invoiced", "Invoices"]
    rows4 = [[f["last"], f["first"], f["email"], f["account"],
              f["last_invoice_date"], round(f["total_invoiced"], 2), f["invoice_count"]]
             for f in sorted(unmatched_tw, key=lambda x: -x["total_invoiced"])]
    write_sheet(wb, "Mismatches", hdr4, rows4)

    Path(args.out).parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.out)

    # ── Step 5 (opt-in): stamp tutor props onto list-3104 contacts ──
    if args.write_props:
        print(f"📤 STEP 5 — writing tutor props to list {GAP_LIST_ID} contacts...")
        prop_rows = [(fam["contact"]["id"], fam["contact"]["email"],
                      f["last_lesson"]["tutor"], f["last_lesson"].get("student_first", ""))
                     for fam, f in gap_with_inv
                     if f.get("last_lesson", {}).get("tutor")]
        write_tutor_props(prop_rows)

    print("\n══ SUMMARY ══")
    print(f"  Charter deals since {SINCE}: {len(deals)}")
    print(f"  Families: {len(families)}  (staff contacts excluded: {len(staff_ids)})")
    print(f"  Tab 1  GAP - Priority Targets:        {len(rows1)}"
          f"  (Hot: {sum(1 for r in rows1 if r[8] == 'Hot')},"
          f" Win-back: {sum(1 for r in rows1 if r[8] == 'Win-back')})")
    print(f"  Tab 2  GAP - Deal But Never Invoiced: {len(rows2)}")
    print(f"  Tab 3  Renewed 26-27:                 {len(rows3)}")
    print(f"  Tab 4  Mismatches (TW no charter):    {len(rows4)}")
    print(f"  Tutor match rate: {with_tutor}/{total_gap} gap families "
          f"({with_tutor / total_gap * 100:.0f}%)")
    print(f"  Saved: {args.out}")


if __name__ == "__main__":
    main()
